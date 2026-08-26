"""
Tratamento de dados das bases baixadas do Looker.

Cada uma das 5 bases tem seu próprio fluxo de tratamento dedicado
(`_process_*`), com regras próprias de seleção de colunas, "período atual"
e destino final - ver `process_base` para o dispatch entre elas.
"""

import logging
import time
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Callable

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.filters import DateGroupItem, FilterColumn, Filters

import config

logger = logging.getLogger("data_processor")

DATE_COLUMN_NUMERO_CONTRATOS = "Dt Relatório"

VERDE_LINHA_NOVA = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMARELO_LINHA_EDITADA = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
VERMELHO_SEM_DADOS = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

HISTORICO_COLUNAS = ["Data/Hora", "Base", "Linhas baixadas", "Linhas novas", "Linhas totais (destino)", "Observação"]

TENTATIVAS_ARQUIVO_BLOQUEADO = 12
ESPERA_ARQUIVO_BLOQUEADO_SEGUNDOS = 10  # total ~2min de espera - o OneDrive pode
                                         # segurar um arquivo bloqueado por mais de
                                         # 15s após gravações rápidas e sucessivas


def _com_retry_arquivo_bloqueado(
    path: Path, acao: Callable[[], object],
    *, tentativas: int = TENTATIVAS_ARQUIVO_BLOQUEADO, espera_segundos: int = ESPERA_ARQUIVO_BLOQUEADO_SEGUNDOS,
) -> object:
    """
    Executa `acao()` - um `DataFrame.to_excel(...)`, `Workbook.save(...)`
    OU `pd.read_excel(...)` - com retry em caso de arquivo bloqueado,
    comum em produção porque as planilhas ficam em pastas do OneDrive
    (arquivo aberto no Excel, ou OneDrive sincronizando). Sem isso, um
    `PermissionError` (WinError 32) derrubava a base inteira mesmo sendo
    uma condição tipicamente temporária.

    Tenta `tentativas` vezes com espera fixa entre elas; na última, deixa o
    erro se propagar com uma mensagem clara sobre a causa provável, em vez
    do `PermissionError` cru do openpyxl/pandas.
    """
    for tentativa in range(1, tentativas + 1):
        try:
            return acao()
        except PermissionError:
            if tentativa == tentativas:
                raise PermissionError(
                    f"Não foi possível acessar '{path}' após {tentativas} tentativas - "
                    "o arquivo parece estar aberto no Excel ou o OneDrive ainda está "
                    "sincronizando. Feche a planilha (ou aguarde a sincronização) e "
                    "rode a base de novo."
                ) from None
            logger.warning(
                "Arquivo '%s' está bloqueado (aberto no Excel ou sincronizando no "
                "OneDrive) - tentativa %d/%d, tentando de novo em %ds...",
                path, tentativa, tentativas, espera_segundos,
            )
            time.sleep(espera_segundos)


def _historico_path() -> Path:
    return config.LOG_DIR / "historico_execucoes.xlsx"


def registrar_historico(
    base_nome: str,
    linhas_baixadas: int | None,
    linhas_novas: int | None = None,
    linhas_total: int | None = None,
    observacao: str = "",
):
    """
    Registra uma linha em `logs/historico_execucoes.xlsx` (separado do
    `rpa.log`, pra abrir/filtrar direto no Excel) com a quantidade de
    linhas baixadas do Looker nesta execução para uma base.

    `linhas_baixadas=0` (ou `None`, quando nem chegou a baixar - ver
    `main.run_bases`) pinta a linha de vermelho. Chamada uma vez por base a
    cada execução, tanto no fluxo de sucesso quanto na falha de download.
    """
    path = _historico_path()
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico"
        ws.append(HISTORICO_COLUNAS)

    ws.append([
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        base_nome,
        linhas_baixadas,
        linhas_novas,
        linhas_total,
        observacao,
    ])

    if not linhas_baixadas:  # 0 ou None
        for cell in ws[ws.max_row]:
            cell.fill = VERMELHO_SEM_DADOS

    ws.auto_filter.ref = ws.dimensions
    _com_retry_arquivo_bloqueado(path, lambda: wb.save(path))


def _chave_como_serie(df: pd.DataFrame, chave) -> pd.Series:
    """Converte a(s) coluna(s) de `chave` numa série de valores comparáveis - uma
    tupla por linha se `chave` for uma lista de colunas, ou a linha inteira
    (como tupla) se `chave` for None.

    Dois ajustes evitam falsos positivos ao comparar com a execução anterior:
    colunas numéricas são arredondadas (6 casas decimais, para absorver a
    perda de precisão do Excel ao salvar/reabrir) e células vazias (NaN) são
    trocadas por um marcador fixo (já que `NaN != NaN` sempre marcaria a
    linha como "diferente").
    """
    if chave is None:
        colunas = df
    elif isinstance(chave, (list, tuple)):
        colunas = df[list(chave)]
    else:
        return df[chave].fillna("__NaN__")
    colunas = colunas.round(6).fillna("__NaN__")
    return colunas.apply(tuple, axis=1)


def _marcar_linhas_novas_e_editadas(path: Path, df: pd.DataFrame, chave, df_anterior: pd.DataFrame | None):
    """Pinta, na planilha "Prévia" salva em `path`, de **verde** as linhas cuja
    chave não existia na versão anterior da Prévia (linhas novas desta
    compilação do RPA) e de **amarelo** as linhas cuja chave já existia mas
    algum dado da linha mudou (linhas editadas) - para o usuário identificar
    visualmente o que foi adicionado/alterado na execução do dia."""
    if df.empty:
        return

    cols_chave = list(chave) if isinstance(chave, (list, tuple)) else ([chave] if chave is not None else [])
    tem_anterior = (
        df_anterior is not None
        and not df_anterior.empty
        and all(c in df_anterior.columns for c in cols_chave)
    )
    if tem_anterior:
        chave_anterior_serie = _chave_como_serie(df_anterior, chave)
        linha_anterior_serie = _chave_como_serie(df_anterior, None)
        linha_anterior_por_chave = dict(zip(chave_anterior_serie, linha_anterior_serie))
    else:
        linha_anterior_por_chave = {}

    chave_serie = _chave_como_serie(df, chave)
    linha_atual_serie = _chave_como_serie(df, None)

    wb = load_workbook(path)
    ws = wb.active
    ws.title = ws.title.lower()  # nome da aba sempre minúsculo, independente de como veio
    mudou_algo = False
    for offset, (k, linha_atual) in enumerate(zip(chave_serie, linha_atual_serie), start=2):  # linha 1 = cabeçalho
        if k not in linha_anterior_por_chave:
            fill = VERDE_LINHA_NOVA
        elif linha_anterior_por_chave[k] != linha_atual:
            fill = AMARELO_LINHA_EDITADA
        else:
            fill = None
        if fill is not None:
            mudou_algo = True
            for cell in ws[offset]:
                cell.fill = fill
    if mudou_algo:
        _com_retry_arquivo_bloqueado(path, lambda: wb.save(path))


def _acumular_planilha(
    caminho: Path, df_novo: pd.DataFrame, chave,
) -> tuple[pd.DataFrame, pd.DataFrame | None]:
    """
    Lê a planilha salva na execução anterior e mescla com os dados novos
    (`df_novo`), em vez de substituir o arquivo por inteiro - garante que
    uma linha já vista nunca desapareça, mesmo se o Looker deixar de
    trazê-la numa execução seguinte. Sempre mantém a versão mais recente de
    cada chave (`drop_duplicates(..., keep="last")`), então uma linha
    existente é atualizada se algum dado dela mudar.

    Se `df_novo` vier vazio, a planilha anterior é devolvida sem alteração.

    Retorna (df a gravar, versão anterior - usada por
    `_marcar_linhas_novas_e_editadas` para a marcação de cor).
    """
    df_anterior = _com_retry_arquivo_bloqueado(caminho, lambda: pd.read_excel(caminho)) if caminho.exists() else None

    if df_novo.empty:
        df_final = df_anterior if df_anterior is not None else df_novo
        return df_final, df_anterior

    if df_anterior is None or df_anterior.empty:
        return df_novo, df_anterior

    df_final = pd.concat([df_anterior, df_novo], ignore_index=True)
    df_final = df_final.drop_duplicates(subset=chave, keep="last")
    return df_final, df_anterior


def _acumular_e_colorir_origem(
    origem_path: Path, df_previa: pd.DataFrame, chave, aplicar_autofiltro: bool,
    *, ordenar: Callable[[pd.DataFrame], pd.DataFrame] | None = None,
) -> tuple[pd.DataFrame, int]:
    """
    Mescla `df_previa` na planilha de origem oficial em `origem_path`,
    usando o mesmo acúmulo "nunca descarta uma linha" (`_acumular_planilha`)
    e a mesma marcação de cor (`_marcar_linhas_novas_e_editadas`): verde
    para chave nova, amarelo para chave existente com dado alterado. Usada
    pelas bases "Dias sem Produção", "Meta Financiamento e Seguro" e
    "Número de Contratos" (Comissão à Vista tem regra própria de origem
    oficial - ver `_acumular_origem_comissao_a_vista`).

    `ordenar`, se informado, reordena o resultado final (ex: por data
    crescente) antes de salvar e colorir - a marcação de cor compara por
    chave, não por posição, então a ordenação não afeta a detecção de
    linha nova/editada.

    Retorna (DataFrame final gravado, quantidade de chaves novas nesta
    execução - usada no histórico).
    """
    df_final, df_anterior = _acumular_planilha(origem_path, df_previa, chave)
    if ordenar is not None:
        df_final = ordenar(df_final)
    _com_retry_arquivo_bloqueado(origem_path, lambda: df_final.to_excel(origem_path, index=False, sheet_name="sheet1"))
    if aplicar_autofiltro:
        _apply_excel_autofilter(origem_path)
    _marcar_linhas_novas_e_editadas(origem_path, df_final, chave, df_anterior)
    return df_final, _contar_chaves_novas(df_final, df_anterior, chave)


def _ordenar_por_data(df: pd.DataFrame) -> pd.DataFrame:
    """Ordena por `DATE_COLUMN_NUMERO_CONTRATOS` ("Dt Relatório") ascendente
    e de forma estável - usada como `ordenar` em `_acumular_e_colorir_origem`
    para a base "numero_contratos" (data real, dia/mês/ano, diferente do
    formato AAAAMM de `_ordenar_por_coluna`)."""
    if DATE_COLUMN_NUMERO_CONTRATOS not in df.columns or df.empty:
        return df
    return df.sort_values(by=DATE_COLUMN_NUMERO_CONTRATOS, ascending=True, kind="stable").reset_index(drop=True)


def _ordenar_por_coluna(coluna: str) -> Callable[[pd.DataFrame], pd.DataFrame]:
    """Fábrica de função `ordenar` para `_acumular_e_colorir_origem`: ordena
    ascendente e de forma estável pela coluna informada (ex: "Anomes
    Apuracao", "Safra Mes" - ambas no formato AAAAMM, que já ordena
    cronologicamente como número/texto puro, sem precisar de parsing de data)."""
    def _ordenar(df: pd.DataFrame) -> pd.DataFrame:
        if coluna not in df.columns or df.empty:
            return df
        return df.sort_values(by=coluna, ascending=True, kind="stable").reset_index(drop=True)
    return _ordenar


def _contar_chaves_novas(df_final: pd.DataFrame, df_anterior: pd.DataFrame | None, chave) -> int:
    """Quantas chaves de `df_final` não existiam em `df_anterior` - usado só
    para o número reportado no histórico (a marcação de cor é feita à parte
    por `_marcar_linhas_novas_e_editadas`)."""
    if df_anterior is not None and not df_anterior.empty:
        chaves_anteriores = set(_chave_como_serie(df_anterior, chave))
        chaves_atuais = set(_chave_como_serie(df_final, chave))
        return len(chaves_atuais - chaves_anteriores)
    return len(df_final)


def _current_month_mask(df: pd.DataFrame, date_col: str) -> pd.Series:
    ano, mes = config.periodo_referencia_atual()
    dt = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)
    return (dt.dt.month == mes) & (dt.dt.year == ano)


def _current_month_mask_com_virada(df: pd.DataFrame, date_col: str, dias_extra: int = 3) -> pd.Series:
    """
    Igual a `_current_month_mask`, mas no primeiro dia do mês (virada)
    também mantém os últimos `dias_extra` dias do mês anterior - alguns
    contratos de fim de mês só aparecem como "PROPOSTA PAGA" com um pequeno
    atraso. A partir do segundo dia do mês, volta a ser só o mês atual.
    """
    hoje = date.today()
    mask = _current_month_mask(df, date_col)

    if hoje.day == 1:
        dt = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)
        primeiro_dia_mes_atual = pd.Timestamp(hoje.year, hoje.month, 1)
        limite_inferior = primeiro_dia_mes_atual - timedelta(days=dias_extra)
        mask = mask | ((dt >= limite_inferior) & (dt < primeiro_dia_mes_atual))

    return mask


def _eh_ano_corrente(ano: int) -> bool:
    """Um ano fechado (ex: 2025) nunca deve ser escrito de novo pela
    automação - só o ano corrente (`date.today().year`) recebe novas linhas.
    Usado ao rotear Comissão à Vista e Dias sem Produção por ano."""
    return ano == date.today().year


def _select_columns(df_novo: pd.DataFrame, base: dict) -> pd.DataFrame:
    """Mantém somente as colunas listadas em `colunas_manter` (config.py) - as demais são excluídas."""
    colunas_manter = base["regras"]["colunas_manter"]
    faltando = [c for c in colunas_manter if c not in df_novo.columns]
    if faltando:
        logger.warning("Colunas esperadas não encontradas no arquivo baixado: %s", faltando)
    colunas_presentes = [c for c in colunas_manter if c in df_novo.columns]
    return df_novo[colunas_presentes]


def _apply_row_filters(df: pd.DataFrame, base: dict) -> pd.DataFrame:
    """Aplica filtros de linha (ex: Status Proposta) antes de qualquer corte de coluna."""
    status_filtro = base["regras"].get("filtro_status_proposta")
    if status_filtro:
        df = df[df["Status Proposta"] == status_filtro]
    return df


def _apply_excel_autofilter(path: Path):
    """Adiciona o filtro (AutoFilter) do Excel em todas as colunas da planilha final."""
    wb = load_workbook(path)
    ws = wb.active
    ws.title = ws.title.lower()  # nome da aba sempre minúsculo, independente de como veio
    ws.auto_filter.ref = ws.dimensions
    _com_retry_arquivo_bloqueado(path, lambda: wb.save(path))


def _mes_atual_e_anterior() -> tuple[tuple[int, int], tuple[int, int]]:
    """(ano, mês) do mês corrente e do mês civil anterior - ex: hoje em
    agosto/2026 retorna ((2026, 8), (2026, 7))."""
    hoje = date.today()
    if hoje.month == 1:
        anterior = (hoje.year - 1, 12)
    else:
        anterior = (hoje.year, hoje.month - 1)
    return (hoje.year, hoje.month), anterior


def _filtrar_mes_atual_e_anterior(
    path: Path, coluna: str, *, is_date: bool = False, coluna_totais: str | None = None,
):
    """
    Aplica no Excel um AutoFilter já ATIVADO na coluna `coluna`, mostrando só
    o mês corrente e o anterior - linhas de outros meses ficam ocultas (não
    apagadas), e a lista de valores do filtro reflete o que está visível,
    pra o Excel já abrir filtrado. Reaplicado do zero a cada execução, então
    a janela rola sozinha com o tempo. Usada nas 5 planilhas oficiais do ano
    corrente, nunca no histórico fechado (2025) nem nas Prévias.

    `is_date=True` para coluna de data real ("Dt Relatório") em vez do
    formato AAAAMM (Anomes Apuracao/Safra Mes/Anomes).
    `coluna_totais`: coluna cuja célula vazia identifica a linha de totais
    (Comissão à Vista), que nunca é ocultada.
    """
    (ano_atual, mes_atual), (ano_anterior, mes_anterior) = _mes_atual_e_anterior()

    wb = load_workbook(path)
    ws = wb.active
    ws.title = ws.title.lower()

    cabecalho = [cell.value for cell in ws[1]]
    if coluna not in cabecalho:
        logger.warning("Coluna '%s' não encontrada em '%s' - filtro de mês não aplicado.", coluna, path)
        return
    col_idx0 = cabecalho.index(coluna)
    col_idx1 = col_idx0 + 1
    idx_totais1 = cabecalho.index(coluna_totais) + 1 if coluna_totais and coluna_totais in cabecalho else None

    valores_visiveis = set()
    for row in range(2, ws.max_row + 1):
        if idx_totais1 is not None and ws.cell(row=row, column=idx_totais1).value is None:
            ws.row_dimensions[row].hidden = False  # linha de totais: sempre visível
            continue

        valor = ws.cell(row=row, column=col_idx1).value
        if valor is None:
            ws.row_dimensions[row].hidden = True
            continue

        if is_date:
            dt = valor if isinstance(valor, (datetime, date)) else pd.to_datetime(valor, dayfirst=True, errors="coerce")
            if pd.isna(dt):
                ws.row_dimensions[row].hidden = True
                continue
            visivel = (dt.year, dt.month) in ((ano_atual, mes_atual), (ano_anterior, mes_anterior))
        else:
            anomes = int(str(int(valor))[:6])
            ano_v, mes_v = anomes // 100, anomes % 100
            visivel = (ano_v, mes_v) in ((ano_atual, mes_atual), (ano_anterior, mes_anterior))
            if visivel:
                valores_visiveis.add(str(valor))

        ws.row_dimensions[row].hidden = not visivel

    ws.auto_filter.ref = ws.dimensions
    ws.auto_filter.filterColumn = []  # remove filtro de execuções anteriores antes de aplicar o atual
    if is_date:
        ws.auto_filter.filterColumn.append(FilterColumn(colId=col_idx0, filters=Filters(dateGroupItem=[
            DateGroupItem(year=ano_atual, month=mes_atual, dateTimeGrouping="month"),
            DateGroupItem(year=ano_anterior, month=mes_anterior, dateTimeGrouping="month"),
        ])))
    else:
        ws.auto_filter.add_filter_column(col_idx0, sorted(valores_visiveis), blank=False)

    _com_retry_arquivo_bloqueado(path, lambda: wb.save(path))
    logger.info(
        "Filtro de mês (%02d/%d e %02d/%d) aplicado em '%s'.",
        mes_anterior, ano_anterior, mes_atual, ano_atual, path,
    )


CHAVE_UNICA_NUMERO_CONTRATOS = "ID Proposta"


def _process_numero_contratos(downloaded_path: Path, base: dict) -> Path:
    """
    Fluxo específico da base "Número de Contratos":
      1. Filtra Status Proposta = PROPOSTA PAGA e seleciona as colunas certas.
      2. Mescla o download inteiro (ano corrente inteiro, "Year To Date") na
         planilha anual "Digitação Analítico - {ano}", via
         `_acumular_e_colorir_origem` - nunca remove uma linha (histórico
         permanente do ano), sempre reordenada por data crescente após o
         merge. Só é escrita para o ano corrente - um ano fechado (ex:
         2025) nunca é tocado de novo, já que `ano` aqui é sempre
         `date.today().year`.
      3. A planilha "Prévia" continua restrita ao mês/ano de referência
         (mês civil atual). Exceção: no primeiro dia do mês, também mantém
         os últimos 3 dias do mês anterior, para não perder contratos de
         fim de mês que aparecem como "PROPOSTA PAGA" com atraso.
         Deduplicação por "ID Proposta", mantendo a versão mais recente.
    """
    chave = CHAVE_UNICA_NUMERO_CONTRATOS
    date_col = DATE_COLUMN_NUMERO_CONTRATOS

    def _apenas_mes_atual(df: pd.DataFrame) -> pd.DataFrame:
        if date_col in df.columns and not df.empty:
            return df[_current_month_mask_com_virada(df, date_col)]
        return df

    df_tratado = pd.read_excel(downloaded_path)
    linhas_baixadas = len(df_tratado)
    df_tratado = _apply_row_filters(df_tratado, base)
    df_tratado = _select_columns(df_tratado, base)  # NÃO recorta por mês - mantém o ano inteiro do download

    # --- 1. Mescla o download inteiro na planilha anual, reordenada por data crescente ---
    ano = date.today().year
    autofiltro = bool(base["regras"].get("aplicar_autofiltro_excel"))
    origem_path = config.caminho_planilha_origem_numero_contratos_anual(ano)
    origem_path.parent.mkdir(parents=True, exist_ok=True)
    df_final, linhas_novas = _acumular_e_colorir_origem(
        origem_path, df_tratado, chave, autofiltro, ordenar=_ordenar_por_data,
    )
    logger.info(
        "Planilha anual atualizada: %s (+%d contratos novos, %d no total)",
        origem_path, linhas_novas, len(df_final),
    )
    _filtrar_mes_atual_e_anterior(origem_path, date_col, is_date=True)

    # --- 2. Acumula na "Prévia" (só o mês/ano de referência), sem duplicar por ID Proposta ---
    df_tratado_mes_atual = _apenas_mes_atual(df_tratado)

    previa_path = config.caminho_previa_numero_contratos()
    previa_path.parent.mkdir(parents=True, exist_ok=True)

    df_previa_existente = _com_retry_arquivo_bloqueado(previa_path, lambda: pd.read_excel(previa_path)) if previa_path.exists() else pd.DataFrame(columns=df_tratado.columns)
    df_previa_existente = _apenas_mes_atual(df_previa_existente)  # descarta sobra de mês anterior já acumulada

    df_previa = pd.concat([df_previa_existente, df_tratado_mes_atual], ignore_index=True)
    df_previa = df_previa.drop_duplicates(subset=chave, keep="last")
    df_previa = _ordenar_por_data(df_previa)

    _com_retry_arquivo_bloqueado(previa_path, lambda: df_previa.to_excel(previa_path, index=False, sheet_name="sheet1"))
    if base["regras"].get("aplicar_autofiltro_excel"):
        _apply_excel_autofilter(previa_path)
    _marcar_linhas_novas_e_editadas(previa_path, df_previa, chave, df_previa_existente)
    logger.info("Prévia atualizada (sem duplicar '%s'): %s (%d linhas)", chave, previa_path, len(df_previa))

    observacao = "Sem dados no período" if linhas_baixadas == 0 else ""
    registrar_historico(base["nome"], linhas_baixadas, linhas_novas, len(df_final), observacao)
    return origem_path


CHAVE_UNICA_DIAS_SEM_PRODUCAO = ["Cd Loja", "Safra Mes"]


def _process_dias_sem_producao(downloaded_path: Path, base: dict) -> Path:
    """
    Fluxo específico da base "Dias sem Produção":
      1. Seleciona as colunas certas (essa base não tem filtro de status).
      2. Acumula o resultado tratado na "Prévia", mesclando sem duplicar
         por `Cd Loja` + `Safra Mes` (ver `_acumular_planilha`) - uma linha
         nunca é descartada, mesmo que o Looker não a traga numa execução
         seguinte.
      3. Mescla essa Prévia em cada planilha de origem oficial por ano
         (nunca remove uma linha - ver `_acumular_e_colorir_origem`),
         roteada a partir do ano em "Safra Mes" - mesmo padrão de
         `_process_meta_financiamento_seguro`. Um ano fechado (ex: 2025)
         nunca é escrito de novo, só o ano corrente (`_eh_ano_corrente`).
    """
    chave = CHAVE_UNICA_DIAS_SEM_PRODUCAO

    df_tratado = pd.read_excel(downloaded_path)
    linhas_baixadas = len(df_tratado)
    df_tratado = _select_columns(df_tratado, base)

    previa_path = config.caminho_previa_dias_sem_producao()
    previa_path.parent.mkdir(parents=True, exist_ok=True)
    df_previa, df_previa_anterior = _acumular_planilha(previa_path, df_tratado, chave)

    _com_retry_arquivo_bloqueado(previa_path, lambda: df_previa.to_excel(previa_path, index=False, sheet_name="sheet1"))
    if base["regras"].get("aplicar_autofiltro_excel"):
        _apply_excel_autofilter(previa_path)
    _marcar_linhas_novas_e_editadas(previa_path, df_previa, chave, df_previa_anterior)
    logger.info("Base '%s' tratada (prévia): %s (%d linhas)", base["nome"], previa_path, len(df_previa))

    if df_previa.empty:
        logger.warning(
            "Relatório '%s' baixado veio vazio e a Prévia também está vazia - "
            "nada para atualizar na planilha de origem oficial.",
            base["nome"],
        )
        registrar_historico(base["nome"], linhas_baixadas, 0, None, "Sem dados no período")
        return previa_path

    autofiltro = bool(base["regras"].get("aplicar_autofiltro_excel"))
    anos_presentes = sorted(df_previa["Safra Mes"].astype(str).str[:4].unique())
    origem_paths = []
    linhas_total_todos_anos = 0
    linhas_novas_todos_anos = 0

    for ano_str in anos_presentes:
        ano = int(ano_str)
        if not _eh_ano_corrente(ano):
            logger.info(
                "Base '%s': ano %d é histórico fechado - não escrevendo de novo na planilha de origem.",
                base["nome"], ano,
            )
            continue

        df_ano_previa = df_previa[df_previa["Safra Mes"].astype(str).str[:4] == ano_str]

        origem_path = config.caminho_planilha_origem_dias_sem_producao(ano)
        origem_path.parent.mkdir(parents=True, exist_ok=True)
        df_final, linhas_novas = _acumular_e_colorir_origem(
            origem_path, df_ano_previa, chave, autofiltro, ordenar=_ordenar_por_coluna("Safra Mes"),
        )

        logger.info(
            "Planilha de origem atualizada: %s (+%d linhas novas, %d no total)",
            origem_path, linhas_novas, len(df_final),
        )
        _filtrar_mes_atual_e_anterior(origem_path, "Safra Mes")
        origem_paths.append(origem_path)
        linhas_total_todos_anos += len(df_final)
        linhas_novas_todos_anos += linhas_novas

    observacao = "Sem dados no período" if linhas_baixadas == 0 else ""
    registrar_historico(base["nome"], linhas_baixadas, linhas_novas_todos_anos, linhas_total_todos_anos, observacao)
    return origem_paths[-1] if origem_paths else previa_path


CHAVE_UNICA_META_FINANCIAMENTO_SEGURO = ["Anomes Apuracao", "Filial"]


def _process_meta_financiamento_seguro(downloaded_path: Path, base: dict) -> Path:
    """
    Fluxo específico da base "Meta Financiamento e Seguro". O download já
    vem com o ANO CORRENTE inteiro (filtro "Safra Mês" -> "is in the year",
    ver `looker_automation.download_meta_financiamento_seguro_report`), não
    só o mês atual - isso garante que qualquer dado adicionado ou editado
    em meses anteriores do próprio ano corrente também seja capturado.

      1. Seleciona as colunas certas (essa base não tem filtro de status).
      2. Acumula o resultado tratado na "Prévia", mesclando sem duplicar por
         `Anomes Apuracao` + `Filial` (ver `_acumular_planilha`) - uma
         linha nunca é descartada, mesmo que o Looker não a traga numa
         execução seguinte.
      3. Mescla essa Prévia em cada planilha de origem oficial (uma por
         ano, nunca remove uma linha, sempre reordenada por "Anomes
         Apuracao" crescente - ver `_acumular_e_colorir_origem`). Roteado
         por ano a partir de TODOS os anos presentes na Prévia acumulada,
         mas só o ano corrente é de fato escrito - um ano fechado (ex:
         2025) nunca é tocado de novo (ver `_eh_ano_corrente`).

    Cada ano tem seu próprio arquivo de origem, sem subpasta:
    "Meta Financiamento Seguro - {ano}.xlsx".
    """
    chave = CHAVE_UNICA_META_FINANCIAMENTO_SEGURO

    df_tratado = pd.read_excel(downloaded_path)
    linhas_baixadas = len(df_tratado)
    df_tratado = _select_columns(df_tratado, base)

    previa_path = config.caminho_previa_meta_financiamento_seguro()
    previa_path.parent.mkdir(parents=True, exist_ok=True)
    df_previa, df_previa_anterior = _acumular_planilha(previa_path, df_tratado, chave)

    _com_retry_arquivo_bloqueado(previa_path, lambda: df_previa.to_excel(previa_path, index=False, sheet_name="sheet1"))
    if base["regras"].get("aplicar_autofiltro_excel"):
        _apply_excel_autofilter(previa_path)
    _marcar_linhas_novas_e_editadas(previa_path, df_previa, chave, df_previa_anterior)
    logger.info("Base '%s' tratada (prévia): %s (%d linhas)", base["nome"], previa_path, len(df_previa))

    if df_previa.empty:
        # Sem "Anomes Apuracao" na Prévia - não dá pra saber qual planilha de
        # origem tocar. Não é erro - loga e segue sem mexer na origem oficial.
        logger.warning(
            "Relatório '%s' baixado veio vazio e a Prévia também está vazia - "
            "nada para atualizar na planilha de origem oficial.",
            base["nome"],
        )
        registrar_historico(base["nome"], linhas_baixadas, 0, None, "Sem dados no período")
        return previa_path

    autofiltro = bool(base["regras"].get("aplicar_autofiltro_excel"))
    anos_presentes = sorted(df_previa["Anomes Apuracao"].astype(str).str[:4].unique())
    origem_paths = []
    linhas_total_todos_anos = 0
    linhas_novas_todos_anos = 0

    for ano_str in anos_presentes:
        ano = int(ano_str)
        if not _eh_ano_corrente(ano):
            logger.info(
                "Base '%s': ano %d é histórico fechado - não escrevendo de novo na planilha de origem.",
                base["nome"], ano,
            )
            continue

        df_ano_previa = df_previa[df_previa["Anomes Apuracao"].astype(str).str[:4] == ano_str]

        origem_path = config.caminho_planilha_origem_meta_financiamento_seguro(ano)
        origem_path.parent.mkdir(parents=True, exist_ok=True)
        df_final, linhas_novas = _acumular_e_colorir_origem(
            origem_path, df_ano_previa, chave, autofiltro, ordenar=_ordenar_por_coluna("Anomes Apuracao"),
        )

        logger.info(
            "Planilha de origem atualizada: %s (+%d linhas novas, %d no total)",
            origem_path, linhas_novas, len(df_final),
        )
        _filtrar_mes_atual_e_anterior(origem_path, "Anomes Apuracao")
        origem_paths.append(origem_path)
        linhas_total_todos_anos += len(df_final)
        linhas_novas_todos_anos += linhas_novas

    observacao = "Sem dados no período" if linhas_baixadas == 0 else ""
    registrar_historico(base["nome"], linhas_baixadas, linhas_novas_todos_anos, linhas_total_todos_anos, observacao)
    return origem_paths[-1] if origem_paths else previa_path


CHAVE_UNICA_CARTEIRA_PARCEIROS = ["Cnpj Da Loja", "Filial", "Anomes"]


def _process_carteira_parceiros(downloaded_path: Path, base: dict) -> Path:
    """
    Fluxo específico da base "Carteira e Parceiros":
      1. Não há filtro de colunas nem de status - o arquivo baixado é usado
         como está (todas as colunas).
      2. Acumula o resultado na "Prévia", mas só o mês atual (mesmo padrão
         de "Número de Contratos"): qualquer linha de "Anomes" diferente do
         mês/ano corrente é descartada antes de mesclar sem duplicar por
         `Cnpj Da Loja` + `Filial` + `Anomes`.
      3. A planilha de origem oficial (uma por ano) continua recebendo o
         download da execução INTEIRO (todos os meses do ano, não só a
         Prévia recortada do passo 2); nunca remove uma linha. Roteado por
         ano a partir de TODOS os anos presentes no download. Um ano já
         fechado (que o filtro "Referência = Este Ano" parou de trazer)
         continua preservado no arquivo de origem daquele ano.

    Em ambas as planilhas, o resultado final fica ordenado por "Anomes"
    crescente (ordenação estável, não embaralha dados dentro do mesmo mês).
    """
    def _ordenar_por_anomes(df: pd.DataFrame) -> pd.DataFrame:
        if "Anomes" in df.columns:
            return df.sort_values(by="Anomes", ascending=True, kind="stable").reset_index(drop=True)
        return df

    def _apenas_mes_atual(df: pd.DataFrame) -> pd.DataFrame:
        if "Anomes" in df.columns and not df.empty:
            ano, mes = config.periodo_referencia_atual()
            anomes_atual = ano * 100 + mes
            return df[df["Anomes"].astype(int) == anomes_atual]
        return df

    df_tratado = pd.read_excel(downloaded_path)
    linhas_baixadas = len(df_tratado)
    df_tratado = _ordenar_por_anomes(df_tratado)

    # --- 1. Acumula na "Prévia", mas só o mês atual (não o ano inteiro) ---
    previa_path = config.caminho_previa_carteira_parceiros()
    previa_path.parent.mkdir(parents=True, exist_ok=True)

    df_previa_existente = _com_retry_arquivo_bloqueado(previa_path, lambda: pd.read_excel(previa_path)) if previa_path.exists() else pd.DataFrame(columns=df_tratado.columns)
    df_previa_existente = _apenas_mes_atual(df_previa_existente)  # descarta sobra de mês anterior já acumulada

    df_previa = pd.concat([df_previa_existente, _apenas_mes_atual(df_tratado)], ignore_index=True)
    df_previa = df_previa.drop_duplicates(subset=CHAVE_UNICA_CARTEIRA_PARCEIROS, keep="last")
    df_previa = _ordenar_por_anomes(df_previa)

    _com_retry_arquivo_bloqueado(previa_path, lambda: df_previa.to_excel(previa_path, index=False, sheet_name="sheet1"))
    if base["regras"].get("aplicar_autofiltro_excel"):
        _apply_excel_autofilter(previa_path)
    _marcar_linhas_novas_e_editadas(previa_path, df_previa, CHAVE_UNICA_CARTEIRA_PARCEIROS, df_previa_existente)
    logger.info("Prévia atualizada (somente mês atual, sem duplicar %s): %s (%d linhas)", CHAVE_UNICA_CARTEIRA_PARCEIROS, previa_path, len(df_previa))

    if df_tratado.empty:
        logger.warning(
            "Relatório '%s' baixado veio vazio - nada para atualizar na planilha de origem oficial.",
            base["nome"],
        )
        registrar_historico(base["nome"], linhas_baixadas, 0, None, "Sem dados no período")
        return previa_path

    # --- 2. Mescla o download INTEIRO (todos os meses do ano) na origem oficial ---
    autofiltro = bool(base["regras"].get("aplicar_autofiltro_excel"))
    anos_presentes = sorted(df_tratado["Anomes"].astype(str).str[:4].unique())
    origem_paths = []
    linhas_total_todos_anos = 0
    linhas_novas_todos_anos = 0

    for ano_str in anos_presentes:
        ano = int(ano_str)
        df_ano_tratado = df_tratado[df_tratado["Anomes"].astype(str).str[:4] == ano_str]

        origem_path = config.caminho_planilha_origem_carteira_parceiros(ano)
        origem_path.parent.mkdir(parents=True, exist_ok=True)
        df_final, df_anterior = _acumular_planilha(origem_path, df_ano_tratado, CHAVE_UNICA_CARTEIRA_PARCEIROS)

        # `pd.concat` (em `_acumular_planilha`) mantém a ordem de colunas da
        # origem já existente - como essa base não tem `colunas_manter` fixo,
        # reordena para bater com a ordem atual do Looker (`df_tratado.columns`)
        # antes de salvar; coluna legada que não existe mais vai para o final.
        colunas_atuais = [c for c in df_tratado.columns if c in df_final.columns]
        colunas_legado = [c for c in df_final.columns if c not in df_tratado.columns]
        df_final = df_final[colunas_atuais + colunas_legado]
        df_final = _ordenar_por_anomes(df_final)

        _com_retry_arquivo_bloqueado(origem_path, lambda: df_final.to_excel(origem_path, index=False, sheet_name="sheet1"))
        if autofiltro:
            _apply_excel_autofilter(origem_path)
        _marcar_linhas_novas_e_editadas(origem_path, df_final, CHAVE_UNICA_CARTEIRA_PARCEIROS, df_anterior)
        linhas_novas = _contar_chaves_novas(df_final, df_anterior, CHAVE_UNICA_CARTEIRA_PARCEIROS)

        logger.info(
            "Planilha de origem atualizada: %s (+%d linhas novas, %d no total)",
            origem_path, linhas_novas, len(df_final),
        )
        if _eh_ano_corrente(ano):  # só a planilha do ano corrente recebe o filtro de mês
            _filtrar_mes_atual_e_anterior(origem_path, "Anomes")
        origem_paths.append(origem_path)
        linhas_total_todos_anos += len(df_final)
        linhas_novas_todos_anos += linhas_novas

    observacao = "Sem dados no período" if linhas_baixadas == 0 else ""
    registrar_historico(base["nome"], linhas_baixadas, linhas_novas_todos_anos, linhas_total_todos_anos, observacao)
    return origem_paths[-1]


def _valor_para_excel(v):
    """Converte um valor de célula pandas/numpy para um tipo que o openpyxl
    aceita em `Worksheet.append` (usado por `_acumular_origem_comissao_a_vista`,
    que grava linha a linha via openpyxl - as demais bases usam
    `DataFrame.to_excel`, que já faz essa conversão sozinho)."""
    if pd.isna(v):
        return None
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v, np.generic):
        v = v.item()
    return v


def _valor_como_texto(v):
    """Converte `v` para string, sem deixar `.0` sobrando quando o valor é um
    float "inteiro" (ex: 202608.0 -> "202608") - usado para colunas que a
    planilha já guarda como texto (ver `_colunas_como_texto`)."""
    if pd.isna(v):
        return None
    if isinstance(v, np.generic):
        v = v.item()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _colunas_como_texto(ws, linha_referencia: int, colunas) -> set:
    """
    Identifica quais colunas a planilha já guarda como texto (não número),
    olhando o tipo do valor em algumas linhas de referência - a planilha
    oficial de "Comissão à Vista" guarda "Cnpj Master", "Cnpj Corban" e
    "Anomes Apuracao" como texto, mas o download do Looker traz esses
    campos como número. Sem converter, CNPJs aparecem em notação
    científica nas linhas novas.

    Confere várias linhas (não só uma) e considera texto se achar QUALQUER
    valor string entre elas, para não deixar uma linha de referência com
    célula vazia mascarar uma coluna que na prática sempre vem como texto.
    """
    colunas_texto = set()
    ultima_linha = min(linha_referencia + 9, ws.max_row)
    for idx, nome in enumerate(colunas, start=1):
        for row in range(linha_referencia, ultima_linha + 1):
            if isinstance(ws.cell(row=row, column=idx).value, str):
                colunas_texto.add(nome)
                break
    return colunas_texto


def _formatar_celula_como(celula, referencia):
    """Copia número/fonte/preenchimento/borda/alinhamento da célula
    `referencia` para `celula` - usado ao anexar linhas na planilha de
    origem oficial de "Comissão à Vista", já que `ws.append()` sozinho não
    herda nenhuma formatação da planilha existente."""
    celula.number_format = referencia.number_format
    celula.font = copy(referencia.font)
    celula.fill = copy(referencia.fill)
    celula.border = copy(referencia.border)
    celula.alignment = copy(referencia.alignment)


def _separar_linha_totais(df: pd.DataFrame) -> tuple[pd.DataFrame, int | None]:
    """
    A planilha de origem oficial de "Comissão à Vista" tem uma linha de
    totais (identificação vazia, soma/média nas colunas numéricas),
    detectada pela coluna "Cd Contrato" vazia. Separa essa linha (se
    existir) do restante dos dados. Retorna (df_sem_totais, posição
    0-based da linha de totais no DataFrame original, ou None se não havia).
    """
    if "Cd Contrato" not in df.columns or df.empty:
        return df, None
    mask_totais = df["Cd Contrato"].isna()
    if not mask_totais.any():
        return df, None
    posicao = int(df.index[mask_totais][0])
    return df[~mask_totais].reset_index(drop=True), posicao


def _calcular_linha_totais(df: pd.DataFrame) -> dict:
    """
    Recalcula a linha de totais da planilha de origem oficial de "Comissão
    à Vista" a partir de TODOS os dados atuais: soma nas colunas "R$...",
    média nas colunas "%...", vazio nas colunas de identificação.
    """
    linha = {}
    for col in df.columns:
        if col.startswith("R$"):
            linha[col] = df[col].sum()
        elif col.startswith("%"):
            linha[col] = df[col].mean()
        else:
            linha[col] = None
    return linha


def _acumular_origem_comissao_a_vista(
    path: Path, df_previa: pd.DataFrame, chave, subset, nome_planilha: str, aplicar_autofiltro: bool,
) -> tuple[int, int]:
    """
    Mescla `df_previa` na planilha de origem oficial de "Comissão à Vista"
    sem duplicar por `chave` - cria a planilha do zero se ainda não
    existir; senão, ACRESCENTA as chaves genuinamente novas (verde) e
    ATUALIZA no lugar as chaves existentes com dado alterado (amarelo) -
    nunca remove uma linha. Editada célula a célula (em vez de
    `DataFrame.to_excel`) para preservar a formatação herdada e a linha de
    totais no final.

    A linha de totais (identificação vazia, soma/média - ver
    `_calcular_linha_totais`) é SEMPRE removida de onde estiver,
    recalculada com todos os dados atuais e recolocada como última linha -
    assim nunca fica presa no meio conforme mais dados são adicionados.

    Retorna (linhas novas + editadas nesta execução, linhas totais depois -
    sem contar a própria linha de totais).

    Em toda gravação (primeira execução ou atualização), o resultado final
    fica ordenado por "Anomes Apuracao" crescente - não só o bloco novo, a
    planilha inteira é reescrita nessa ordem a cada execução.
    """
    coluna_data = "Anomes Apuracao"

    if not path.exists():
        df_previa_ordenada = _ordenar_por_coluna(coluna_data)(df_previa)
        _com_retry_arquivo_bloqueado(path, lambda: df_previa_ordenada.to_excel(path, index=False, sheet_name="sheet1"))
        wb = load_workbook(path)
        ws = wb.active
        ws.title = ws.title.lower()  # nome da aba sempre minúsculo, independente de como veio
        if len(df_previa_ordenada) >= 1:
            totais = _calcular_linha_totais(df_previa_ordenada)
            ws.append([_valor_para_excel(totais.get(c)) for c in df_previa_ordenada.columns])
            linha_referencia = ws.cell(row=2, column=1).row  # 1ª linha de dado real
            for col in range(1, ws.max_column + 1):
                _formatar_celula_como(
                    ws.cell(row=ws.max_row, column=col), ws.cell(row=linha_referencia, column=col),
                )
        _com_retry_arquivo_bloqueado(path, lambda: wb.save(path))
        if aplicar_autofiltro:
            _apply_excel_autofilter(path)
        _marcar_linhas_novas_e_editadas(path, df_previa_ordenada, chave, None)
        logger.info("Planilha '%s' criada (primeira execução): %s (%d linhas)", nome_planilha, path, len(df_previa_ordenada))
        return len(df_previa_ordenada), len(df_previa_ordenada)

    df_existente_completo = _com_retry_arquivo_bloqueado(path, lambda: pd.read_excel(path))
    df_existente, posicao_totais = _separar_linha_totais(df_existente_completo)
    linha_totais_excel = posicao_totais + 2 if posicao_totais is not None else None  # +2: 0-based -> 1-based, +1 pelo cabeçalho

    faltando = [c for c in df_existente.columns if c not in df_previa.columns]
    extras = [c for c in df_previa.columns if c not in df_existente.columns]
    if faltando or extras:
        logger.warning(
            "Colunas da Prévia divergem da planilha '%s' existente "
            "(faltando=%s, extras=%s) - alinhando pelas colunas já existentes na planilha.",
            nome_planilha, faltando, extras,
        )
    df_previa_alinhada = df_previa.reindex(columns=df_existente.columns)
    df_previa_alinhada = df_previa_alinhada.drop_duplicates(subset=subset, keep="last").reset_index(drop=True)

    if not df_existente.empty:
        chave_existente_serie = _chave_como_serie(df_existente, chave)
        linha_existente_serie = _chave_como_serie(df_existente, None)
        posicao_por_chave = {k: i for i, k in enumerate(chave_existente_serie)}
        linha_existente_por_chave = dict(zip(chave_existente_serie, linha_existente_serie))
    else:
        posicao_por_chave = {}
        linha_existente_por_chave = {}

    chave_previa_serie = _chave_como_serie(df_previa_alinhada, chave)
    linha_previa_serie = _chave_como_serie(df_previa_alinhada, None)

    # Separa em "genuinamente nova" (chave nunca vista) e "editada" (chave
    # já existe, mas algum dado mudou) - uma chave existente sem nenhuma
    # mudança não entra em nenhum dos dois grupos (fica como está, sem cor).
    indices_novos = []
    edicoes = []  # (posição em df_existente, índice em df_previa_alinhada)
    for i, (k, linha_atual) in enumerate(zip(chave_previa_serie, linha_previa_serie)):
        if k not in posicao_por_chave:
            indices_novos.append(i)
        elif linha_existente_por_chave[k] != linha_atual:
            edicoes.append((posicao_por_chave[k], i))

    df_realmente_novo = df_previa_alinhada.iloc[indices_novos]

    if df_realmente_novo.empty and not edicoes and linha_totais_excel is None:
        logger.info(
            "Nenhum registro novo/alterado para '%s' nesta execução - planilha mantida como está (%d linhas).",
            nome_planilha, len(df_existente),
        )
        return 0, len(df_existente)

    wb = load_workbook(path)
    ws = wb.active
    ws.title = ws.title.lower()  # nome da aba sempre minúsculo, independente de como veio

    # Referência de formatação: sempre a 1ª linha de dado real (linha 2) -
    # continua válida mesmo depois de remover a linha de totais de
    # qualquer posição (deletar uma linha mais abaixo não afeta a linha 2).
    linha_referencia = 2 if ws.max_row >= 2 else None
    colunas_texto = _colunas_como_texto(ws, linha_referencia, df_existente.columns) if linha_referencia else set()

    # Captura a formatação da linha de referência ANTES de apagar as linhas
    # de dado (o rewrite abaixo reconstrói a planilha inteira em ordem
    # cronológica, então a linha 2 física não existe mais depois do
    # `delete_rows` - sem essa captura, perderíamos a formatação herdada).
    formato_colunas = None
    if linha_referencia:
        formato_colunas = [
            (
                ws.cell(row=linha_referencia, column=col).number_format,
                copy(ws.cell(row=linha_referencia, column=col).font),
                copy(ws.cell(row=linha_referencia, column=col).border),
                copy(ws.cell(row=linha_referencia, column=col).alignment),
            )
            for col in range(1, ws.max_column + 1)
        ]

    def _aplicar_formato_capturado(row_idx: int):
        if not formato_colunas:
            return
        for col, (number_format, font, border, alignment) in enumerate(formato_colunas, start=1):
            cell = ws.cell(row=row_idx, column=col)
            cell.number_format = number_format
            cell.font = font
            cell.border = border
            cell.alignment = alignment

    # Aplica as edições nos dados (ainda em pandas, sem tocar no Excel) e
    # identifica por CHAVE quem é novo/editado - a planilha é reconstruída
    # do zero em ordem cronológica logo abaixo, então cor por posição não
    # serve mais aqui.
    df_existente_atualizada = df_existente.copy()
    for posicao_existente, idx_novo in edicoes:
        df_existente_atualizada.iloc[posicao_existente] = df_previa_alinhada.iloc[idx_novo].values

    chaves_novas = set(chave_previa_serie.iloc[indices_novos])
    chaves_editadas = set(chave_previa_serie.iloc[[idx for _, idx in edicoes]])

    df_tudo = pd.concat([df_existente_atualizada, df_realmente_novo], ignore_index=True)
    df_tudo = _ordenar_por_coluna(coluna_data)(df_tudo)
    chave_tudo_serie = _chave_como_serie(df_tudo, chave)

    sem_fill = PatternFill(fill_type=None)

    # Apaga TODAS as linhas de dado (mantém só o cabeçalho) para reescrever
    # em ordem cronológica - diferente das outras bases (que reescrevem via
    # `DataFrame.to_excel`), aqui precisa ser explícito porque a planilha é
    # editada célula a célula pra preservar a formatação herdada.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    for idx, (_, linha) in enumerate(df_tudo.iterrows()):
        valores = [
            _valor_como_texto(v) if nome in colunas_texto else _valor_para_excel(v)
            for nome, v in linha.items()
        ]
        ws.append(valores)
        row_idx = ws.max_row
        _aplicar_formato_capturado(row_idx)
        chave_linha = chave_tudo_serie.iloc[idx]
        if chave_linha in chaves_novas:
            fill = VERDE_LINHA_NOVA
        elif chave_linha in chaves_editadas:
            fill = AMARELO_LINHA_EDITADA
        else:
            fill = sem_fill
        for cell in ws[row_idx]:
            cell.fill = fill

    totais = _calcular_linha_totais(df_tudo)
    ws.append([_valor_para_excel(totais.get(c)) for c in df_existente.columns])
    _aplicar_formato_capturado(ws.max_row)
    for cell in ws[ws.max_row]:
        cell.fill = sem_fill  # linha de totais nunca é colorida

    _com_retry_arquivo_bloqueado(path, lambda: wb.save(path))
    if aplicar_autofiltro:
        _apply_excel_autofilter(path)

    total = len(df_tudo)
    logger.info(
        "Planilha '%s' atualizada: +%d linhas novas, %d linhas editadas (%d no total, linha de totais recalculada e mantida no final)",
        nome_planilha, len(df_realmente_novo), len(edicoes), total,
    )
    return len(df_realmente_novo) + len(edicoes), total


def _process_comissao_a_vista(downloaded_path: Path, base: dict) -> Path:
    """
    Fluxo específico da base "Comissão à Vista - Analítico": duas
    planilhas são atualizadas a cada execução, com regras diferentes:
      - Prévia: reescrita por inteiro a cada execução, mesclando com o que
        já existia (chave existente é ATUALIZADA, não só ignorada) e
        recalculando as cores do zero - mesmo padrão das outras bases
        (`_marcar_linhas_novas_e_editadas`).
      - Planilha de origem oficial, uma por ano (ex: "Comissão à Vista -
        Analítico - 2026.xlsx"), roteada a partir do ano em "Anomes
        Apuracao": recebe a Prévia já acumulada, filtrada para aquele ano -
        nunca remove uma linha (`_acumular_origem_comissao_a_vista`). Um
        ano fechado (ex: 2025) nunca é escrito de novo, só o ano corrente
        (`_eh_ano_corrente`).
    Chave em ambas: `regras["chave_comparacao"]`, config.py.

    O arquivo baixado traz uma linha de totais no final (colunas de
    identificação vazias, só os valores em R$ somados) - descartada antes
    de qualquer comparação, filtrando por "Cd Contrato" vazio.

    Pelo menos uma coluna do relatório vem do Looker com espaços/quebra de
    linha ao redor do nome, enquanto a planilha de origem oficial não tem
    - por isso os nomes de coluna do download são sempre limpos
    (`str.strip()`) antes de qualquer comparação de colunas.
    """
    previa_path = config.caminho_previa_comissao_a_vista()
    previa_path.parent.mkdir(parents=True, exist_ok=True)
    chave = base["regras"].get("chave_comparacao")
    subset = list(chave) if isinstance(chave, (list, tuple)) else ([chave] if chave else None)
    autofiltro = bool(base["regras"].get("aplicar_autofiltro_excel"))

    df_novo = pd.read_excel(downloaded_path)
    df_novo.columns = df_novo.columns.str.strip()
    if "Cd Contrato" in df_novo.columns:
        linha_totais = df_novo["Cd Contrato"].isna()
        if linha_totais.any():
            logger.info(
                "Removendo %d linha(s) de totais/resumo (sem 'Cd Contrato') do relatório 'Comissão à Vista' baixado.",
                int(linha_totais.sum()),
            )
            df_novo = df_novo[~linha_totais].reset_index(drop=True)
    linhas_baixadas = len(df_novo)
    if df_novo.empty:
        logger.warning("Relatório 'Comissão à Vista' baixado veio vazio - nada para adicionar.")
        registrar_historico(base["nome"], linhas_baixadas, 0, None, "Sem dados no período")
        return previa_path

    # --- Prévia: reescrita + cor recalculada a cada execução (igual as outras 4 bases) ---
    df_previa_anterior = _com_retry_arquivo_bloqueado(previa_path, lambda: pd.read_excel(previa_path)) if previa_path.exists() else None
    if df_previa_anterior is not None and not df_previa_anterior.empty:
        faltando = [c for c in df_previa_anterior.columns if c not in df_novo.columns]
        extras = [c for c in df_novo.columns if c not in df_previa_anterior.columns]
        if faltando or extras:
            logger.warning(
                "Colunas do relatório baixado divergem da Prévia 'Comissão à Vista' existente "
                "(faltando=%s, extras=%s) - alinhando pelas colunas já existentes na planilha.",
                faltando, extras,
            )
        df_novo_alinhado = df_novo.reindex(columns=df_previa_anterior.columns)
        df_previa_final = pd.concat([df_previa_anterior, df_novo_alinhado], ignore_index=True)
        # mantém sempre a versão mais recente baixada de cada chave (permite detectar "editada")
        df_previa_final = df_previa_final.drop_duplicates(subset=subset, keep="last")
    else:
        df_previa_final = df_novo

    _com_retry_arquivo_bloqueado(previa_path, lambda: df_previa_final.to_excel(previa_path, index=False, sheet_name="sheet1"))
    if autofiltro:
        _apply_excel_autofilter(previa_path)
    _marcar_linhas_novas_e_editadas(previa_path, df_previa_final, chave, df_previa_anterior)
    logger.info(
        "Planilha 'Comissão à Vista - Analítico (Prévia)' atualizada: %s (%d linhas)",
        previa_path, len(df_previa_final),
    )

    # --- Planilha de origem oficial: mescla a Prévia já acumulada (nunca perde linha),
    # roteada por ano a partir de "Anomes Apuracao" - só o ano corrente é escrito. ---
    anos_presentes = sorted(df_previa_final["Anomes Apuracao"].astype(str).str[:4].unique())
    origem_paths = []
    linhas_total_todos_anos = 0
    linhas_novas_todos_anos = 0

    for ano_str in anos_presentes:
        ano = int(ano_str)
        if not _eh_ano_corrente(ano):
            logger.info(
                "Base '%s': ano %d é histórico fechado - não escrevendo de novo na planilha de origem.",
                base["nome"], ano,
            )
            continue

        df_ano_previa = df_previa_final[df_previa_final["Anomes Apuracao"].astype(str).str[:4] == ano_str]
        origem_path = config.caminho_planilha_origem_comissao_a_vista(ano)
        origem_path.parent.mkdir(parents=True, exist_ok=True)
        linhas_novas_origem, linhas_total_origem = _acumular_origem_comissao_a_vista(
            origem_path, df_ano_previa, chave, subset, f"Comissão à Vista - Analítico - {ano} (origem oficial)",
            aplicar_autofiltro=autofiltro,
        )
        _filtrar_mes_atual_e_anterior(origem_path, "Anomes Apuracao", coluna_totais="Cd Contrato")
        origem_paths.append(origem_path)
        linhas_total_todos_anos += linhas_total_origem
        linhas_novas_todos_anos += linhas_novas_origem

    observacao = "Sem dados no período" if linhas_baixadas == 0 else ""
    registrar_historico(base["nome"], linhas_baixadas, linhas_novas_todos_anos, linhas_total_todos_anos, observacao)
    return origem_paths[-1] if origem_paths else previa_path


def process_base(downloaded_path: Path, base: dict) -> Path:
    """Pipeline completo para uma base: despacha para o fluxo dedicado conforme o 'modo' configurado."""
    modo = base["regras"].get("modo")

    if modo == "planilha_origem_local":
        return _process_numero_contratos(downloaded_path, base)

    if modo == "planilha_origem_local_dias_sem_producao":
        return _process_dias_sem_producao(downloaded_path, base)

    if modo == "planilha_origem_local_meta_financiamento_seguro":
        return _process_meta_financiamento_seguro(downloaded_path, base)

    if modo == "planilha_origem_local_carteira_parceiros":
        return _process_carteira_parceiros(downloaded_path, base)

    if modo == "planilha_origem_local_comissao_a_vista":
        return _process_comissao_a_vista(downloaded_path, base)

    raise ValueError(f"Base '{base['id']}' não tem 'modo' de tratamento reconhecido em config.py")


def processar_ano_fechado(downloaded_path: Path, base_id: str, ano: int) -> Path:
    """
    Processa a carga única de um ano fechado (ex: 2025) para uma das 4
    bases em `looker_automation.BASES_COM_ANO_FECHADO`, gravando direto na
    planilha de origem oficial daquele ano (verde para tudo, já que é a
    primeira carga) - usada só pelo script de backfill de histórico
    (`backfill_ano_fechado.py`), nunca pelo fluxo diário/mensal normal
    (`process_base`), que por isso nunca reescreve um ano já fechado (ver
    `_eh_ano_corrente`).
    """
    base = config.get_base_by_id(base_id)
    autofiltro = bool(base["regras"].get("aplicar_autofiltro_excel"))

    if base_id == "meta_financiamento_seguro":
        df = pd.read_excel(downloaded_path)
        df = _select_columns(df, base)
        origem_path = config.caminho_planilha_origem_meta_financiamento_seguro(ano)
        origem_path.parent.mkdir(parents=True, exist_ok=True)
        df_final, linhas_novas = _acumular_e_colorir_origem(
            origem_path, df, CHAVE_UNICA_META_FINANCIAMENTO_SEGURO, autofiltro,
            ordenar=_ordenar_por_coluna("Anomes Apuracao"),
        )
        total = len(df_final)

    elif base_id == "dias_sem_producao":
        df = pd.read_excel(downloaded_path)
        df = _select_columns(df, base)
        origem_path = config.caminho_planilha_origem_dias_sem_producao(ano)
        origem_path.parent.mkdir(parents=True, exist_ok=True)
        df_final, linhas_novas = _acumular_e_colorir_origem(
            origem_path, df, CHAVE_UNICA_DIAS_SEM_PRODUCAO, autofiltro,
            ordenar=_ordenar_por_coluna("Safra Mes"),
        )
        total = len(df_final)

    elif base_id == "comissao_a_vista":
        df = pd.read_excel(downloaded_path)
        df.columns = df.columns.str.strip()
        if "Cd Contrato" in df.columns:
            df = df[~df["Cd Contrato"].isna()].reset_index(drop=True)
        chave = base["regras"].get("chave_comparacao")
        subset = list(chave) if isinstance(chave, (list, tuple)) else ([chave] if chave else None)
        origem_path = config.caminho_planilha_origem_comissao_a_vista(ano)
        origem_path.parent.mkdir(parents=True, exist_ok=True)
        linhas_novas, total = _acumular_origem_comissao_a_vista(
            origem_path, df, chave, subset, f"Comissão à Vista - Analítico - {ano} (origem oficial)",
            aplicar_autofiltro=autofiltro,
        )

    elif base_id == "numero_contratos":
        df = pd.read_excel(downloaded_path)
        df = _apply_row_filters(df, base)
        df = _select_columns(df, base)
        origem_path = config.caminho_planilha_origem_numero_contratos_anual(ano)
        origem_path.parent.mkdir(parents=True, exist_ok=True)
        df_final, linhas_novas = _acumular_e_colorir_origem(
            origem_path, df, CHAVE_UNICA_NUMERO_CONTRATOS, autofiltro, ordenar=_ordenar_por_data,
        )
        total = len(df_final)

    else:
        raise ValueError(f"Base '{base_id}' não suporta carga de ano fechado")

    logger.info(
        "Carga do ano fechado %d concluída para '%s': %s (+%d linhas, %d no total)",
        ano, base["nome"], origem_path, linhas_novas, total,
    )
    registrar_historico(f"{base['nome']} - {ano} (ano fechado)", len(df), linhas_novas, total)
    return origem_path
